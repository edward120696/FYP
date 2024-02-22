import PySimpleGUI as sg
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from readability.readability import Document
from requests_html import HTMLSession
import re
from datetime import datetime

"""
    Demo Program - Multithreaded Long Tasks GUI
    
    Presents one method for running long-running operations in a PySimpleGUI environment.
    
    The PySimpleGUI code, and thus the underlying GUI framework, runs as the primary, main thread
    The "long work" is contained in the thread that is being started.
    So that you don't have to import and understand the threading module, this program uses window.start_thread to run a thread.
    
    The thread is using TUPLES for its keys.  This enables you to easily find the thread events by looking at event[0].
        The Thread Keys look something like this:  ('-THREAD-', message)
        If event [0] == '-THREAD-' then you know it's one of these tuple keys.
         
    Copyright 2022 PySimpleGUI"""
    
def process_file(file_path,session,window):
    filepath = file_path        # This is file location.

    wb = load_workbook(filepath)    # Load excel and find active sheet.
    sheet = wb.active

    i = 1                           # This is line number of the first url existing.
    count=1
    countDone=0
    
    clean = re.compile('<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});')
    while i > 0:
        # Take urls one by one
        i = i + 1
        url = sheet.cell(i,1).value
        if url == None:             # If url doesn't exist, iteration is finished
            break
        
        print("No: "+str(count))
        url = ('http://' + url, url) ['http://' in url]         # remake url
        print(url)
        count = count + 1        
        content = ''

        # Take urls one by one. If request state is not 200 (success), then go to the next url
        try:
            page = session.get(url) # call session and use Get method
            page.html.render(sleep=2,timeout=20) # render the page
            document = Document(page.text) 
            links = page.html.absolute_links # Get all urls
            soup = BeautifulSoup(page.text , "html.parser") # parse the HTML
        
            innerhtml = soup.find_all(['p','span']) # use BeautifulSoup to find out all the Tag "a" and "span" content 
            txt = ''
            for idx in innerhtml:       # parse innerhtml
                txt += '\n' + idx.text
            phones = re.findall(r'\(?[0-9]{2,3}\)?[ .-]?[0-9]{3,4}[ .-]?[0-9]{4}', txt)    
            emails = re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', txt)
            address = re.findall(r"\d+\/?[A-Z]?,?\s?(?:[\w\s]+,)?\s?(?:[#\w\-]+,)?\s?[\w\s]+\,?\s?(?:[\w\s]+\,)?\s?(?:[\w\s]+\,)?\s?\w{2,}(?:\s\d{5})?(?:\-\d{4})?(?:\s[A-Za-z]+)?",txt)
            title = document.title()
            link_contact = ''
            link_about = ''
            for link in links:
                if "contact" in link: # find contact page url
                    link_contact = link
                if "about" in link: # find about page url
                    link_about = link
            
            if link_about != '':
                print(link_about) 
                page = session.get(link_about) # find about-us page
                page.html.render(sleep=2,timeout=20)    
                document = Document(page.text) 
                content = document.summary() # find the company industry information

            if link_contact == '' :
                url += '/contact'
            else:
                url = link_contact
            print(url)

            try:
                page = session.get(url) # Find contact page and process above actions again
                page.html.render(sleep=2,timeout=20)
                soup = BeautifulSoup(page.text, "html.parser")
                innerhtml = soup.find_all(['p','span'])
                txt = ''
                for idx in innerhtml:
                    txt += '\n' + idx.text
                phones = re.findall(r'\(?[0-9]{2,3}\)?[ .-]?[0-9]{3,4}[ .-]?[0-9]{4}', txt)    
                emails = re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', txt)
                address = re.findall(r"\d+\/?[A-Z]?,?\s?(?:[\w\s]+,)?\s?(?:[#\w\-]+,)?\s?[\w\s]+\,?\s?(?:[\w\s]+\,)?\s?(?:[\w\s]+\,)?\s?\w{2,}(?:\s\d{5})?(?:\-\d{4})?(?:\s[A-Za-z]+)?",txt)
            except Exception as e:
                print(e)
                continue

        except Exception as e:
            print(e)
            continue
        if phones:
            phones = ', '.join(list(dict.fromkeys(phones)))     # remove duplicate phone numbers            
            sheet.cell(i, 3).value = phones
            print(phones)
        if emails:
            emails = ', '.join(list(dict.fromkeys(emails)))  # remove duplicate emails numbers  
            sheet.cell(i, 5).value = emails 
            print(emails)
        if address:
            address = re.sub(clean,'',str(address))
            sheet.cell(i, 7).value = address   
            print(address)                        
        if title:
            title = re.sub(clean,'',title)
            sheet.cell(i, 9).value = title  
            print(title)
        if content != '':
            content = re.sub(clean,'',content)
            sheet.cell(i, 11).value = content  
            print(content)
        wb.save(filepath)
    session.close()
    now = datetime.now()
    current = now.strftime("%H:%M:%S")
    print("End time = ",current)
    window.write_event_value(('-THREAD-', 'web srcaping end'), 'Done!')  
     
def the_gui():
    """
    Starts and executes the GUI
    Reads data from a Queue and displays the data to the window
    Returns when the user exits / closes the window
    """
    sg.theme('Light Brown 3')

    layout = [[sg.Text('Result output')],
              [sg.Output(size=(70, 12))],
              [sg.Text('Select an Excel file to process:')],
              [sg.Input(key='-FILE-', enable_events=True), sg.FileBrowse()],
              [sg.Button('Process'), sg.Button('Exit')]      
            ]
        
    window = sg.Window('Web Scraping System', layout)

    # --------------------- EVENT LOOP ---------------------
    while True:
        event, values = window.read()
        if event in (sg.WIN_CLOSED, 'Exit'):
            break
        elif event == 'Process':
            print('Now start web scraping...!')
            file_path = values['-FILE-']
            session = HTMLSession()
            session.browser
            now = datetime.now()
            current = now.strftime("%H:%M:%S")
            print("Start time = ",current)
            window.start_thread(lambda: process_file(file_path,session,window), ('-THREAD-', '-THEAD ENDED-'))
        elif event[0] == '-CRPINT-':
            sg.cprint(event[1], colors=event[2])
        elif event[0] == '-THREAD-':
            print('Got a message back from the thread: ', event[1])

    # if user exits the window, then close the window and exit the GUI func
    window.close()

if __name__ == '__main__':
    the_gui()
    print('Exiting Program')